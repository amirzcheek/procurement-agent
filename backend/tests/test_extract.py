"""Тест извлечения xlsx: шапка не в первой строке + объединённые ячейки."""
import io

import openpyxl
import pytest

import extract


def _make_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "КП"
    # Строки-мусор сверху (шапка таблицы не в первой строке).
    ws["A1"] = "Коммерческое предложение"
    ws.merge_cells("A1:E1")  # объединённая ячейка-заголовок
    ws["A2"] = "ТОО Поставщик"
    # Шапка таблицы в 4-й строке.
    ws.append([])  # row 3 пустая
    ws.append(["№", "Наименование", "Кол-во", "Ед.", "Цена"])  # row 4
    ws.append([1, "Бумага А4 SvetoCopy 500л", 10, "пачка", 1850])  # row 5
    ws.append([2, "Картридж HP CF259A", 3, "шт", 42000])  # row 6
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_extract_xlsx_finds_rows():
    res = extract.extract("kp.xlsx", _make_xlsx())
    assert res.source_type == "xlsx"
    assert "Бумага А4 SvetoCopy 500л" in res.text
    assert "Картридж HP CF259A" in res.text
    assert "Коммерческое предложение" in res.text  # merged-заголовок развёрнут


def test_unsupported_format():
    with pytest.raises(extract.ExtractionError):
        extract.extract("file.docx", b"123")


def test_pdf_without_text_raises_ocr(monkeypatch):
    # PDF без текстового слоя + OCR выключен → понятная ошибка NeedsOCRError.
    import config

    monkeypatch.setattr(config.get_settings(), "ocr_enabled", False)
    monkeypatch.setattr(extract, "_pdf_text", lambda content: ("", 1))
    with pytest.raises(extract.NeedsOCRError):
        extract.extract("scan.pdf", b"%PDF-1.4 fake")
