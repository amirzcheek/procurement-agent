"""Тест экспорта xlsx — отчёт строится и открывается openpyxl."""
import io

import openpyxl

import compare
import report as report_mod
from models import AnalysisReport, Item, MatchDecision, NormalizedQuery, PriceHit


def _report():
    item = Item(name="Бумага А4", qty=10, unit="пачка", unit_price=2000)
    hit = PriceHit(price=1800, currency="KZT", title="Бумага А4", url="https://satu.kz/1",
                   source="satu.kz", available=True)
    ir = compare.build_item_report(
        item, NormalizedQuery(query="бумага а4"),
        [(hit, MatchDecision(is_match=True, confidence=0.9, reason="ok"))],
    )
    rep = AnalysisReport(job_id="abc123", filename="kp.xlsx", items=[ir])
    rep.summary = compare.build_summary([ir])
    return rep


def test_build_xlsx_opens():
    data = report_mod.build_xlsx(_report())
    assert isinstance(data, bytes) and len(data) > 0
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Позиции" in wb.sheetnames
    assert "Сводка" in wb.sheetnames
    ws = wb["Позиции"]
    # Шапка + 1 позиция.
    assert ws.max_row >= 2
    assert ws.cell(row=1, column=2).value == "Наименование (КП)"
