"""Извлечение данных из документа: xlsx, текстовый PDF, СКАНЫ (OCR через Gemini).

Каскад:
- xlsx: openpyxl (как раньше).
- PDF: сначала pdfplumber; если текст пустой ИЛИ подозрительно куцый (< OCR_MIN_CHARS_PER_PAGE
  на страницу) → считаем сканом и идём в OCR (рендер страниц в изображения через PyMuPDF).
- изображения (png/jpg/…): сразу OCR.

OCR — Gemini vision (те же реквизиты, что для grounded-поиска). Режим OCR_MODE:
  text (по умолчанию): скан → распознанный текст → parse_items (Qwen структурирует);
  structured: Gemini сразу возвращает позиции JSON (минуя parse_items).

Выход: ExtractResult(text, source_type[xlsx|pdf_text|ocr], items|None).
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from typing import List, Optional

from config import get_settings
from logging_conf import get_logger

log = get_logger("extract")

_IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp")
_MIME = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
         ".tif": "image/tiff", ".tiff": "image/tiff", ".bmp": "image/bmp", ".webp": "image/webp"}


class ExtractionError(Exception):
    """Понятная пользователю ошибка извлечения."""


class NeedsOCRError(ExtractionError):
    """Скан при выключенном OCR."""


@dataclass
class ExtractResult:
    text: str
    source_type: str                 # xlsx | pdf_text | ocr
    items: Optional[List[dict]] = None  # заполняется только в OCR_MODE=structured


def _cell_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ── xlsx (без изменений) ─────────────────────────────────────────────────────
def extract_xlsx(content: bytes) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    out: List[str] = []
    for ws in wb.worksheets:
        merged_ranges = list(ws.merged_cells.ranges)
        fill = {}
        for rng in merged_ranges:
            top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    fill[(r, c)] = top_left
        rows_text: List[str] = []
        for row in ws.iter_rows():
            cells = []
            any_value = False
            for cell in row:
                val = fill.get((cell.row, cell.column), cell.value)
                txt = _cell_text(val)
                if txt:
                    any_value = True
                cells.append(txt)
            if any_value:
                while cells and cells[-1] == "":
                    cells.pop()
                rows_text.append("\t".join(cells))
        if rows_text:
            out.append(f"### Лист: {ws.title}")
            out.extend(rows_text)
            out.append("")
    text = "\n".join(out).strip()
    if not text:
        raise ExtractionError("В xlsx-файле не найдено непустых ячеек.")
    log.info("xlsx: извлечено %d строк текста", text.count("\n") + 1)
    return text


# ── PDF: текстовый слой через pdfplumber ─────────────────────────────────────
def _pdf_text(content: bytes):
    """(текст, число_страниц). Не бросает NeedsOCR — решение принимает вызывающий."""
    import pdfplumber

    parts: List[str] = []
    pages = 0
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        pages = len(pdf.pages)
        for pi, page in enumerate(pdf.pages, start=1):
            page_chunks: List[str] = []
            try:
                tables = page.extract_tables() or []
            except Exception as e:  # pragma: no cover
                log.warning("pdfplumber: ошибка extract_tables стр.%d: %s", pi, e)
                tables = []
            for table in tables:
                for trow in table:
                    cells = [_cell_text(c) for c in trow]
                    if any(cells):
                        page_chunks.append("\t".join(cells))
            txt = page.extract_text() or ""
            if txt.strip():
                page_chunks.append(txt.strip())
            if page_chunks:
                parts.append(f"### Страница {pi}")
                parts.extend(page_chunks)
                parts.append("")
    return "\n".join(parts).strip(), pages


# ── Рендер PDF-страниц в изображения (для OCR) ──────────────────────────────
def _render_pdf_pages(content: bytes, dpi: int) -> List[bytes]:
    import fitz  # PyMuPDF

    images: List[bytes] = []
    with fitz.open(stream=content, filetype="pdf") as doc:
        for page in doc:
            pix = page.get_pixmap(dpi=dpi)
            images.append(pix.tobytes("png"))
    return images


# ── OCR-ветка (Gemini vision) ────────────────────────────────────────────────
def _ocr(images: List[bytes], mimes: List[str]) -> ExtractResult:
    s = get_settings()
    if not s.ocr_enabled:
        raise NeedsOCRError(
            "Это скан (нет текстового слоя), а распознавание сканов отключено (OCR_ENABLED=false). "
            "Загрузите текстовый PDF/xlsx или включите OCR."
        )
    if (s.ocr_provider or "gemini").lower() != "gemini":
        raise ExtractionError(f"OCR_PROVIDER={s.ocr_provider} не поддерживается (доступен gemini).")

    import gemini
    if not gemini.is_configured():
        raise ExtractionError("OCR требует GEMINI_API_KEY (тот же, что для поиска) — он не задан.")

    structured = (s.ocr_mode or "text").lower() == "structured"
    log.info("OCR: страниц=%d, режим=%s", len(images), "structured" if structured else "text")

    try:
        if structured:
            items: List[dict] = []
            for img, mime in zip(images, mimes):
                items.extend(gemini.ocr_page_items(img, mime))
            if not items:
                raise ExtractionError("Gemini не распознал позиции на скане (пустой результат).")
            # текст для отладки/лога — краткая сводка
            text = "\n".join(str(it.get("name", "")) for it in items)
            return ExtractResult(text=text, source_type="ocr", items=items)

        texts: List[str] = []
        for pi, (img, mime) in enumerate(zip(images, mimes), start=1):
            page_text = gemini.ocr_page_text(img, mime)
            if page_text.strip():
                texts.append(f"### Страница {pi}\n{page_text.strip()}")
        text = "\n\n".join(texts).strip()
        if not text:
            raise ExtractionError("Gemini вернул пустой результат распознавания скана.")
        log.info("OCR: распознано %d символов", len(text))
        return ExtractResult(text=text, source_type="ocr")
    except ExtractionError:
        raise
    except Exception as e:
        log.warning("OCR через Gemini упал: %s", e)
        raise ExtractionError(f"Не удалось распознать скан (Gemini недоступен): {e}")


# ── Диспетчер ────────────────────────────────────────────────────────────────
def extract(filename: str, content: bytes) -> ExtractResult:
    """Каскад извлечения. Возвращает ExtractResult."""
    name = (filename or "").lower()
    s = get_settings()

    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return ExtractResult(text=extract_xlsx(content), source_type="xlsx")

    if name.endswith(_IMAGE_EXTS):
        ext = "." + name.rsplit(".", 1)[-1]
        return _ocr([content], [_MIME.get(ext, "image/png")])

    if name.endswith(".pdf"):
        text, pages = _pdf_text(content)
        threshold = max(1, pages) * s.ocr_min_chars_per_page
        if text and len(text) >= threshold:
            log.info("pdf(text): %d символов на %d стр.", len(text), pages)
            return ExtractResult(text=text, source_type="pdf_text")
        # пусто или подозрительно куцо → скан → OCR
        if not s.ocr_enabled:
            raise NeedsOCRError(
                "Это скан (нет текстового слоя), а распознавание сканов отключено "
                "(OCR_ENABLED=false). Загрузите текстовый PDF/xlsx или включите OCR."
            )
        log.info("pdf: текст куцый (%d < %d) — считаю сканом, OCR", len(text), threshold)
        images = _render_pdf_pages(content, s.ocr_dpi)
        if not images:
            raise ExtractionError("Не удалось отрендерить страницы PDF для OCR.")
        return _ocr(images, ["image/png"] * len(images))

    if name.endswith(".xls"):
        raise ExtractionError("Старый формат .xls не поддерживается. Сохраните как .xlsx.")
    raise ExtractionError(
        f"Неподдерживаемый формат файла: {filename}. Нужен .xlsx, .pdf или изображение (png/jpg)."
    )
