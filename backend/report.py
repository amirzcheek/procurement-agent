"""Этап 8. Экспорт отчёта в xlsx (openpyxl): лист позиций + лист сводки, заливка по флагам."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models import AnalysisReport

_FLAG_FILL = {
    "green": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "yellow": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "red": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "gray": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}
_FLAG_RU = {"green": "Норма", "yellow": "Внимание", "red": "Завышение", "gray": "Проверить"}
_HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


_RISK_FILL = {
    "low": _FLAG_FILL["green"],
    "medium": _FLAG_FILL["yellow"],
    "high": _FLAG_FILL["red"],
    "unknown": _FLAG_FILL["gray"],
}
_RISK_RU = {"low": "Низкий", "medium": "Средний", "high": "Высокий", "unknown": "Нет данных"}


def _add_historical_sheet(wb: Workbook, historical: dict) -> None:
    """Лист исторического анализа цен за выбранный период (min/max по двум источникам)."""
    period = historical.get("period_label", "")
    ws = wb.create_sheet("История цен")
    ws.append([f"Исторический анализ цен — {period}"])
    ws["A1"].font = Font(bold=True)
    headers = [
        "Наименование", "Цена КП", "Внутр. мин", "Внутр. макс", "Внутр. n",
        "Веб мин", "Веб макс", "Веб n", "Диапазон мин", "Диапазон макс",
        "Тренд (веб) %", "Риск", "Рекомендация",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for it in historical.get("items", []):
        internal = it.get("internal") or {}
        web = it.get("web") or {}
        risk = it.get("risk_level", "unknown")
        ws.append([
            it.get("name", ""),
            it.get("kp_unit_price"),
            internal.get("min"), internal.get("max"), internal.get("count"),
            web.get("min"), web.get("max"), web.get("count"),
            it.get("combined_min"), it.get("combined_max"),
            web.get("trend_pct"),
            _RISK_RU.get(risk, risk),
            it.get("recommendation") or it.get("message") or "",
        ])
        ws.cell(row=ws.max_row, column=12).fill = _RISK_FILL.get(risk, _RISK_FILL["unknown"])

    widths = [42, 12, 12, 12, 8, 12, 12, 8, 12, 12, 11, 12, 46]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A3"


_FACTOR_SEP = "; "


def build_conclusion_xlsx(conc: dict) -> bytes:
    """Заключение по договору → xlsx: листы «Заключение», «Проверки», «История цен»."""
    wb = Workbook()
    c = conc.get("contract", {})
    risk = conc.get("risk_level", "unknown")

    # ── Лист «Заключение» ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Заключение"
    rows = [
        ["ЗАКЛЮЧЕНИЕ ПО ДОГОВОРУ", ""],
        ["Номер", c.get("number")],
        ["Дата", c.get("date")],
        ["Поставщик", c.get("supplier")],
        ["Заказчик", c.get("customer")],
        ["Предмет", c.get("subject")],
        ["Сумма", c.get("total_sum")],
        ["Источник финансирования", c.get("funding_source")],
        ["Статус", c.get("status")],
        ["Период сравнения цен", conc.get("period_label")],
        ["ИТОГОВЫЙ РИСК", _RISK_RU.get(risk, risk)],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=13)
    ws.cell(row=len(rows), column=2).fill = _RISK_FILL.get(risk, _RISK_FILL["unknown"])
    ws.cell(row=len(rows), column=1).font = Font(bold=True)

    ws.append([])
    ws.append(["Факторы риска", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for f in conc.get("risk_factors", []):
        detail = f.get("item") or (_FACTOR_SEP.join(f["detail"]) if isinstance(f.get("detail"), list) else f.get("detail") or "")
        ws.append([f.get("factor"), f"{detail}  [{f.get('source')}]"])
    ws.append([])
    ws.append(["Рекомендации", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for rec in conc.get("recommendations", []):
        ws.append(["•", rec])
    ws.append([])
    ws.append(["ДИСКЛЕЙМЕР", conc.get("disclaimer", "")])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80
    for row in ws.iter_rows(min_col=2, max_col=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # ── Лист «Проверки» ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Проверки")
    ws2.append(["Проверка", "Риск", "Итог/находки"])
    for col in range(1, 4):
        ws2.cell(row=1, column=col).fill = _HEADER_FILL
        ws2.cell(row=1, column=col).font = _HEADER_FONT
    _names = {"characteristics": "Характеристики", "price": "Цена",
              "quantity": "Количество", "conditions": "Обязательные условия"}
    for chk in conc.get("checks", []):
        f = chk.get("findings") or {}
        summary = _friendly_findings(chk.get("type"), f, chk.get("result") or {})
        ws2.append([_names.get(chk["type"], chk["type"]), _RISK_RU.get(chk["risk_level"], chk["risk_level"]), summary])
        ws2.cell(row=ws2.max_row, column=2).fill = _RISK_FILL.get(chk["risk_level"], _RISK_FILL["unknown"])
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 80
    for row in ws2.iter_rows(min_col=3, max_col=3):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # ── Лист «История цен» (за период) ───────────────────────────────────────
    _add_historical_sheet(wb, {"period_label": conc.get("period_label", ""), "items": conc.get("items", [])})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _friendly_findings(ctype: str, f: dict, r: dict) -> str:
    if ctype == "conditions":
        parts = []
        if f.get("missing_critical"):
            parts.append("нет: " + ", ".join(f["missing_critical"]))
        if f.get("missing_extra"):
            parts.append("нет доп.: " + ", ".join(f["missing_extra"]))
        return "; ".join(parts) or "все условия присутствуют"
    if ctype == "quantity":
        ms = f.get("mismatches") or []
        return r.get("note") or (f"расхождения: {len(ms)}" if ms else "совпадает")
    if ctype == "characteristics":
        items = f.get("items") or []
        if not items:
            return r.get("note") or "нет данных для сравнения"
        return "; ".join(f"{it['item']}: {', '.join(it.get('differences') or [])}" for it in items[:5])
    if ctype == "price":
        high = f.get("high") or []
        return f"позиций с высоким риском: {len(high)}" if high else "в пределах диапазона"
    return ""


def build_xlsx(report: AnalysisReport, historical: dict | None = None) -> bytes:
    wb = Workbook()

    # ── Лист «Позиции» ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Позиции"
    headers = [
        "№", "Наименование (КП)", "Кол-во", "Ед.", "Цена КП",
        "Мин рынок", "Медиана", "Макс рынок", "Дельта %", "Флаг",
        "Ср. confidence", "Переплата (оц.)", "Подтв. цен", "Ссылки", "Комментарий",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    for i, r in enumerate(report.items, start=1):
        links = " | ".join(c.url for c in r.confirmed_prices[:5])
        ws.append([
            i,
            r.item.name,
            r.item.qty,
            r.item.unit,
            round(r.kp_unit_price, 2) if r.kp_unit_price is not None else None,
            round(r.market_min, 2) if r.market_min is not None else None,
            round(r.market_median, 2) if r.market_median is not None else None,
            round(r.market_max, 2) if r.market_max is not None else None,
            round(r.delta_pct, 1) if r.delta_pct is not None else None,
            _FLAG_RU.get(r.flag, r.flag),
            round(r.avg_confidence, 2) if r.avg_confidence is not None else None,
            round(r.estimated_overpay, 2) if r.estimated_overpay is not None else None,
            len(r.confirmed_prices),
            links,
            r.flag_reason or (r.error or ""),
        ])
        # Заливка строки по флагу (колонка «Флаг»).
        ws.cell(row=i + 1, column=10).fill = _FLAG_FILL.get(r.flag, _FLAG_FILL["gray"])

    widths = [4, 45, 7, 8, 12, 12, 12, 12, 9, 12, 12, 14, 9, 50, 40]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"

    # ── Лист «Сводка» ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Сводка")
    s = report.summary
    rows = [
        ["Файл", report.filename],
        ["Job ID", report.job_id],
        ["Всего позиций", s.total_items],
        ["🟢 Норма (green)", s.green],
        ["🟡 Внимание (yellow)", s.yellow],
        ["🔴 Завышение (red)", s.red],
        ["⚪ Проверить (gray)", s.gray],
        ["Оценочная переплата, " + s.currency, round(s.estimated_total_overpay, 2)],
        ["", ""],
        ["ДИСКЛЕЙМЕР", report.disclaimer],
    ]
    for row in rows:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 70
    ws2["B10"].alignment = Alignment(wrap_text=True, vertical="top")
    for rkey, flag in (("A4", "green"), ("A5", "yellow"), ("A6", "red"), ("A7", "gray")):
        ws2[rkey].fill = _FLAG_FILL[flag]

    # ── Лист «История цен» (Этап 1) ─────────────────────────────────────────
    if historical and historical.get("enabled") and historical.get("items"):
        _add_historical_sheet(wb, historical)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
